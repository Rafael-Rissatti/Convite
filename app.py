import os
from datetime import datetime
from io import BytesIO
from xml.sax.saxutils import escape

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

app = Flask(__name__)
app.config['SECRET_KEY'] = 'pokemon_secret_key'
app.config['SQLALCHEMY_DATABASE_PATH'] = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'rsvp.db')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + app.config['SQLALCHEMY_DATABASE_PATH']
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

class RSVP(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    acompanhante = db.Column(db.String(100), nullable=True)
    status = db.Column(db.String(10), nullable=False) # 'sim' ou 'nao'

with app.app_context():
    db.create_all()

ADMIN_PASSWORD = 'pika'

def admin_required():
    return session.get('logged_in')

def get_rsvp_rows():
    respostas = RSVP.query.order_by(RSVP.status.desc(), RSVP.nome.asc()).all()
    return [
        {
            "id": resposta.id,
            "nome": resposta.nome,
            "acompanhante": resposta.acompanhante or "-",
            "status": "Confirmado" if resposta.status == "sim" else "Ausente",
        }
        for resposta in respostas
    ]

def safe_excel_text(value):
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    nome = request.form.get('nome')
    acompanhante = request.form.get('acompanhante')
    status = request.form.get('status')
    
    if nome and status:
        novo_rsvp = RSVP(nome=nome, acompanhante=acompanhante, status=status)
        db.session.add(novo_rsvp)
        db.session.commit()
        return jsonify({"status": "success", "message": "Temos que pegar todos! Sua presença foi registrada com sucesso!"})
    return jsonify({"status": "error", "message": "Por favor, preencha todos os campos obrigatórios."}), 400

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('admin'))
        return render_template('login.html', error="Senha da Equipe Rocket? Tente novamente!")
    return render_template('login.html')

@app.route('/admin')
def admin():
    if not admin_required():
        return redirect(url_for('login'))
    
    confirmados = RSVP.query.filter_by(status='sim').all()
    ausentes = RSVP.query.filter_by(status='nao').all()
    
    return render_template('admin.html', confirmados=confirmados, ausentes=ausentes)

@app.route('/api/stats')
def stats():
    if not admin_required():
        return jsonify({"error": "Unauthorized"}), 401
        
    count_sim = RSVP.query.filter_by(status='sim').count()
    count_nao = RSVP.query.filter_by(status='nao').count()
    
    return jsonify({
        "sim": count_sim,
        "nao": count_nao
    })

@app.route('/admin/export/excel')
def export_excel():
    if not admin_required():
        return redirect(url_for('login'))

    rows = get_rsvp_rows()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Respostas"

    headers = ["ID", "Treinador", "Acompanhante", "Status"]
    sheet.append(headers)

    for row in rows:
        sheet.append([
            row["id"],
            safe_excel_text(row["nome"]),
            safe_excel_text(row["acompanhante"]),
            row["status"],
        ])

    header_fill = PatternFill("solid", fgColor="3B4CCA")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)

    summary = workbook.create_sheet("Resumo")
    confirmados = sum(1 for row in rows if row["status"] == "Confirmado")
    ausentes = sum(1 for row in rows if row["status"] == "Ausente")
    summary.append(["Resumo da Batalha", ""])
    summary.append(["Confirmados", confirmados])
    summary.append(["Ausentes", ausentes])
    summary.append(["Total de Respostas", len(rows)])
    summary["A1"].font = Font(bold=True, color="3B4CCA", size=14)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 16

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="relatorio_respostas_convite.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route('/admin/export/pdf')
def export_pdf():
    if not admin_required():
        return redirect(url_for('login'))

    rows = get_rsvp_rows()
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=28,
        leftMargin=28,
        topMargin=28,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Relatório de Respostas do Convite", styles["Title"]),
        Paragraph(datetime.now().strftime("Gerado em %d/%m/%Y às %H:%M"), styles["Normal"]),
        Spacer(1, 14),
    ]

    table_data = [["ID", "Treinador", "Acompanhante", "Status"]]
    for row in rows:
        table_data.append([
            row["id"],
            Paragraph(escape(row["nome"]), styles["BodyText"]),
            Paragraph(escape(row["acompanhante"]), styles["BodyText"]),
            row["status"],
        ])

    table = Table(table_data, colWidths=[42, 250, 250, 110], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B4CCA")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D7D9E4")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF7DA")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))

    elements.append(table)
    document.build(elements)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="relatorio_respostas_convite.pdf",
        mimetype="application/pdf",
    )

@app.route('/admin/delete/<int:id>', methods=['POST'])
def delete_rsvp(id):
    if not admin_required():
        return jsonify({"error": "Unauthorized"}), 401
    
    rsvp = RSVP.query.get_or_404(id)
    db.session.delete(rsvp)
    db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/edit/<int:id>', methods=['POST'])
def edit_rsvp(id):
    if not admin_required():
        return jsonify({"error": "Unauthorized"}), 401
    
    rsvp = RSVP.query.get_or_404(id)
    rsvp.nome = request.form.get('nome')
    rsvp.acompanhante = request.form.get('acompanhante')
    rsvp.status = request.form.get('status')
    
    db.session.commit()
    return redirect(url_for('admin'))

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
